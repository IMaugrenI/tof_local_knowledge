from __future__ import annotations

import csv
import hashlib
import json
import mimetypes
from email import policy
from email.parser import BytesParser
from pathlib import Path
from typing import Any

from bs4 import BeautifulSoup
from docx import Document as DocxDocument
from openpyxl import load_workbook
from pypdf import PdfReader


TEXT_SUFFIXES = {'.txt', '.md', '.log', '.py', '.rst', '.ini', '.cfg', '.yaml', '.yml'}


def safe_resolve_under(base: str | Path, candidate: str | Path) -> Path:
    base_path = Path(base).resolve()
    candidate_path = Path(candidate).resolve()
    candidate_path.relative_to(base_path)
    return candidate_path


def sha256_file(path: str | Path) -> str:
    digest = hashlib.sha256()
    with open(path, 'rb') as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b''):
            digest.update(chunk)
    return digest.hexdigest()


def mime_type_for(path: str | Path) -> str:
    guessed, _ = mimetypes.guess_type(str(path))
    return guessed or 'application/octet-stream'


def text_block(kind: str, text: str, page_no: int | None = None) -> dict[str, Any]:
    return {'kind': kind, 'page_no': page_no, 'text': text}


def extract_blocks(path: str | Path) -> tuple[list[dict[str, Any]], list[str]]:
    file_path = Path(path)
    suffix = file_path.suffix.lower()
    warnings: list[str] = []

    try:
        if suffix in TEXT_SUFFIXES:
            return [text_block('text_block', file_path.read_text(encoding='utf-8', errors='replace'))], warnings

        if suffix == '.json':
            data = json.loads(file_path.read_text(encoding='utf-8', errors='replace'))
            return [text_block('json', json.dumps(data, ensure_ascii=False, indent=2))], warnings

        if suffix in {'.html', '.htm'}:
            soup = BeautifulSoup(file_path.read_text(encoding='utf-8', errors='replace'), 'html.parser')
            return [text_block('html', soup.get_text('\n', strip=True))], warnings

        if suffix == '.csv':
            rows: list[dict[str, Any]] = []
            with open(file_path, 'r', encoding='utf-8', errors='replace', newline='') as handle:
                reader = csv.reader(handle)
                for idx, row in enumerate(reader, start=1):
                    rows.append(text_block('csv_row', ' | '.join(row), idx))
            return rows, warnings

        if suffix == '.eml':
            with open(file_path, 'rb') as handle:
                message = BytesParser(policy=policy.default).parse(handle)
            parts = [
                f"Subject: {message.get('subject', '')}",
                f"From: {message.get('from', '')}",
                f"To: {message.get('to', '')}",
            ]
            body = message.get_body(preferencelist=('plain', 'html'))
            if body:
                parts.append(body.get_content())
            return [text_block('email', '\n'.join(parts))], warnings

        if suffix == '.pdf':
            reader = PdfReader(str(file_path))
            blocks: list[dict[str, Any]] = []
            for idx, page in enumerate(reader.pages, start=1):
                text = (page.extract_text() or '').strip()
                if text:
                    blocks.append(text_block('pdf_page', text, idx))
                else:
                    warnings.append(f'empty_pdf_page:{idx}')
            return blocks, warnings

        if suffix == '.docx':
            doc = DocxDocument(str(file_path))
            paragraphs = [p.text.strip() for p in doc.paragraphs if p.text and p.text.strip()]
            return [text_block('docx_paragraph', text, idx) for idx, text in enumerate(paragraphs, start=1)], warnings

        if suffix == '.xlsx':
            workbook = load_workbook(str(file_path), read_only=True, data_only=True)
            blocks: list[dict[str, Any]] = []
            for sheet in workbook.worksheets:
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    values = [str(cell) for cell in row if cell is not None and str(cell).strip()]
                    if values:
                        blocks.append(text_block(f'xlsx_row:{sheet.title}', ' | '.join(values), row_idx))
            return blocks, warnings

        warnings.append(f'unsupported_suffix:{suffix or "none"}')
        return [], warnings
    except Exception as exc:  # pragma: no cover - defensive path
        return [], [f'extraction_error:{type(exc).__name__}']
